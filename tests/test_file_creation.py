import os
import pytest
from openpyxl import load_workbook
from utils.tax_calculator import calculate_and_save  # Импортируем из основного модуля
from faker import Faker


@pytest.fixture
def files_dir():
    """Фикстура для тестовой директории."""
    test_dir = os.path.dirname(os.path.abspath(__file__))
    output_folder = os.path.join(test_dir, "files")

    # Создаем директорию если не существует
    os.makedirs(output_folder, exist_ok=True)

    file_path = os.path.join(output_folder, "net_salaries.xlsx")

    # Удаляем файл перед тестом, если существует
    if os.path.exists(file_path):
        os.remove(file_path)

    yield file_path

    # Очищаем после теста
    if os.path.exists(file_path):
        os.remove(file_path)


def test_excel_file_creation(files_dir):
    """Тестирование создания и содержимого Excel файла."""
    fake = Faker()
    test_gross_salary = fake.random_int(min=50000, max=300000)
    test_start_month = fake.random_int(min=0, max=11)

    calculate_and_save(test_gross_salary, test_start_month, files_dir)

    assert os.path.exists(files_dir), f"Файл не был создан по пути: {files_dir}"

    # Проверяем содержимое файла
    wb = load_workbook(files_dir)
    sheet = wb.active

    assert sheet['A1'].value == 'Месяц', "Неверный заголовок столбца месяцев"
    assert sheet['B1'].value == 'Зарплата на руки', "Неверный заголовок столбца зарплат"
    assert sheet.max_row > 1, "Файл не содержит данных"

    for row in range(2, sheet.max_row + 1):
        assert isinstance(sheet[f'A{row}'].value, str), f"Месяц в строке {row} должен быть строкой"
        assert sheet[f'B{row}'].value.endswith(" руб."), f"Зарплата в строке {row} должна быть в рублях"
